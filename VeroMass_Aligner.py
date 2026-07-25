#!/usr/bin/env python3
"""
VeroMass Aligner  v1.8.0
VeroMass / MoleculeID Platform — STANDALONE UTILITY

PURPOSE : Point at a folder of LC-MS run files (mzML / mzXML / Thermo RAW /
          MGF, and .zip archives of any of these) and produce a cross-sample
          aligned feature table — chromatographic peak detection ->
          correspondence grouping -> retention time correction -> re-grouping
          -> Excel export, with per-sample peak detail and representative MS2
          fragmentation attached per feature.

INPUTS : .mzML / .mzXML / .raw / .mgf, plus .zip archives (extracted
          recursively — nested zips supported). MGF spectra are already
          peak-picked, so that path skips detection entirely (fastest).

PERFORMANCE : peak detection is fully vectorized (flatten -> sort-by-m/z ->
          gap-split ROIs); a ~2,300-scan file that took ~90s in v1.1 now runs
          in ~1-2s. RAW->mzML conversions are pre-run concurrently.

PERMANENTLY STANDALONE: no shared code, database, config, or output with
MGF Extractor, MoleculeID Processor, or Phyto CrossMatcher.

Author  : VeroMass / MoleculeID team
Date    : 2026-07-20
"""
import os
import re
import sys
import queue
import bisect
import base64
import struct
import zipfile
import threading
import subprocess
import time
import datetime
import tempfile
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext, messagebox

# ═════════════════════════════════════════════════════════════════════════════
# Auto-install runtime dependencies
# ═════════════════════════════════════════════════════════════════════════════
def _ensure_deps():
    import importlib
    for pkg in ["numpy", "scipy", "pandas", "openpyxl", "requests"]:
        try:
            importlib.import_module(pkg)
        except ImportError:
            flags = [sys.executable, "-m", "pip", "install", pkg, "-q"]
            if sys.platform != "win32":
                flags.insert(-1, "--break-system-packages")
            subprocess.check_call(flags, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

_ensure_deps()

import numpy as np
import pandas as pd
import requests as req
from scipy.signal import find_peaks
from scipy.ndimage import gaussian_filter1d
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ═════════════════════════════════════════════════════════════════════════════
# Constants
# ═════════════════════════════════════════════════════════════════════════════
TOOL_NAME     = "VeroMass Aligner"
VERSION       = "1.9.0"
OUTPUT_SUBDIR = "VeroMass_Aligner_Output"
MS_EXTS = (".mzml", ".mzxml", ".raw", ".mgf")   # directly-readable MS files
ARCHIVE_EXTS = (".zip",)                          # extracted, then scanned for MS_EXTS
SUPPORTED_EXTS = MS_EXTS + ARCHIVE_EXTS           # what the folder scanner picks up

# ── Colour palette (matches MoleculeID design system) ─────────────────────────
C_BG     = "#0C1120"
C_PANEL  = "#141C2E"
C_BORDER = "#243050"
C_FG     = "#E8EFF8"
C_TEAL   = "#00E5C0"
C_AMB    = "#FFB940"
C_PURP   = "#9B8DF8"
C_GREEN  = "#36D47E"
C_RED    = "#FF7B6B"
C_DIM    = "#4A5878"
C_SUB    = "#8A9BBB"
C_MONO   = "#4FAAFF"


# ═════════════════════════════════════════════════════════════════════════════
# LOW-LEVEL BINARY / RT HELPERS  (shared convention with MoleculeID_Processor)
# ═════════════════════════════════════════════════════════════════════════════
def _decode_binary(b64, is_zlib=True, prec=32, big_end=False):
    import zlib
    try:
        data = base64.b64decode(b64.strip())
        if is_zlib:
            try:
                data = zlib.decompress(data)
            except Exception:
                pass
        n = len(data) // (4 if prec == 32 else 8)
        if n == 0:
            return []
        end = ">" if big_end else "<"
        fmt = f"{end}{n}{'f' if prec == 32 else 'd'}"
        return list(struct.unpack(fmt, data[: n * (4 if prec == 32 else 8)]))
    except Exception:
        return []


def _rt_to_min(s):
    s = str(s).strip()
    m = re.search(r"PT(\d+)M([\d.]+)S", s, re.I)
    if m:
        return int(m.group(1)) + float(m.group(2)) / 60
    m = re.search(r"PT?([\d.]+)S", s, re.I)
    if m:
        return float(m.group(1)) / 60
    try:
        return float(s) / 60
    except Exception:
        return 0.0


def _cluster_frags(pairs, tol=0.02):
    """Cluster near-identical fragment ions within tol Da; returns (mz, max_intensity) per cluster."""
    if not pairs:
        return []
    sp = sorted(pairs, key=lambda x: x[0])
    clusters = []
    cur_mzs = [sp[0][0]]
    cur_ints = [sp[0][1]]
    for mz, inten in sp[1:]:
        if mz - cur_mzs[0] <= tol:
            cur_mzs.append(mz)
            cur_ints.append(inten)
        else:
            total_i = sum(cur_ints)
            avg_mz = sum(m * i for m, i in zip(cur_mzs, cur_ints)) / total_i
            clusters.append((avg_mz, max(cur_ints)))
            cur_mzs, cur_ints = [mz], [inten]
    total_i = sum(cur_ints)
    avg_mz = sum(m * i for m, i in zip(cur_mzs, cur_ints)) / total_i
    clusters.append((avg_mz, max(cur_ints)))
    return clusters


def summarize_fragments(pairs, n=20):
    """Top-N fragment ions as a semicolon-delimited, base-peak-normalised string."""
    try:
        clean = [(float(a), float(b)) for a, b in pairs if float(b) > 0]
    except Exception:
        return ""
    if not clean:
        return ""
    clustered = _cluster_frags(clean, tol=0.02)
    if not clustered:
        return ""
    max_c = max(b for _, b in clustered)
    if not max_c:
        return ""
    return "; ".join(
        f"{a:.4f}({b / max_c * 100:.0f}%)"
        for a, b in sorted(clustered, key=lambda x: -x[1])[:n]
    )


# ═════════════════════════════════════════════════════════════════════════════
# MS1 + MS2 FULL-SCAN READERS  (adapted from MoleculeID_Processor's
# parse_mzml_full / parse_mzxml_full — extended to keep full centroid
# m/z+intensity arrays per MS1 scan, needed for chromatographic peak
# detection, not just TIC/base-peak; MS2 scans are kept for fragment lookup.)
# ═════════════════════════════════════════════════════════════════════════════
def read_ms_scans_mzml(path, cb=None):
    """Returns (ms1_scans, ms2_scans).
    ms1_scans: [{'rt_min','mz':ndarray,'inten':ndarray}]
    ms2_scans: [{'rt_min','prec_mz','pairs':[(mz,inten),...]}]
    """
    import xml.etree.ElementTree as ET
    NS = "http://psi.hupo.org/ms/mzml"

    def tag(e):
        return e.tag.split("}")[-1]

    def cv(elem, acc):
        for cp in elem.iter(f"{{{NS}}}cvParam"):
            if cp.get("accession") == acc:
                return cp.get("value")
        return None

    ms1, ms2 = [], []
    n = 0
    for _event, elem in ET.iterparse(path, events=("end",)):
        if tag(elem) != "spectrum":
            continue
        n += 1
        if cb and n % 2000 == 0:
            cb(f"  Parsed {n:,} scans…")
        lvl = int(cv(elem, "MS:1000511") or 1)
        rt_min = 0.0
        for cp in elem.iter(f"{{{NS}}}cvParam"):
            acc = cp.get("accession", "")
            if acc in ("MS:1000016", "MS:1000896"):
                try:
                    rv = float(cp.get("value", 0))
                    unit = cp.get("unitAccession", "")
                    rt_min = rv / 60 if (unit == "UO:0000010" or acc == "MS:1000896") else rv
                except Exception:
                    pass
                break
        mz_arr = int_arr = []
        for bda in elem.iter(f"{{{NS}}}binaryDataArray"):
            is_mz = is_int = is_zlib = False
            prec = 32
            for cp in bda.iter(f"{{{NS}}}cvParam"):
                acc = cp.get("accession", "")
                if acc == "MS:1000514":
                    is_mz = True
                elif acc == "MS:1000515":
                    is_int = True
                elif acc == "MS:1000574":
                    is_zlib = True
                elif acc == "MS:1000523":
                    prec = 64
            bin_el = next(iter(bda.iter(f"{{{NS}}}binary")), None)
            if bin_el is None:
                bin_el = next(iter(bda.iter("binary")), None)
            if bin_el is not None and bin_el.text:
                arr = _decode_binary(bin_el.text, is_zlib, prec, False)
                if is_mz:
                    mz_arr = arr
                elif is_int:
                    int_arr = arr
        if lvl == 1:
            if mz_arr and int_arr:
                mz_np = np.asarray(mz_arr, dtype=np.float64)
                in_np = np.asarray(int_arr, dtype=np.float64)
                order = np.argsort(mz_np)
                ms1.append({"rt_min": rt_min, "mz": mz_np[order], "inten": in_np[order]})
        elif lvl == 2:
            prec_mz = 0.0
            for prec_tag in elem.iter(f"{{{NS}}}precursor"):
                for cp in prec_tag.iter(f"{{{NS}}}cvParam"):
                    if cp.get("accession") == "MS:1000744":
                        try:
                            prec_mz = float(cp.get("value", 0))
                        except Exception:
                            pass
                break
            pairs = sorted(zip(mz_arr, int_arr), key=lambda x: x[0]) if mz_arr and int_arr else []
            ms2.append({"rt_min": rt_min, "prec_mz": prec_mz, "pairs": pairs})
        elem.clear()
    ms1.sort(key=lambda s: s["rt_min"])
    ms2.sort(key=lambda s: s["rt_min"])
    return ms1, ms2


def read_ms_scans_mzxml(path, cb=None):
    import xml.etree.ElementTree as ET

    ms1, ms2 = [], []
    n = 0
    for _event, elem in ET.iterparse(path, events=("end",)):
        if elem.tag.split("}")[-1] != "scan":
            continue
        n += 1
        if cb and n % 2000 == 0:
            cb(f"  Parsed {n:,} scans…")
        lvl = int(elem.get("msLevel", "1"))
        rt_min = _rt_to_min(elem.get("retentionTime", "0"))
        peaks_el = elem.find("peaks")
        mz_arr = int_arr = []
        if peaks_el is not None and peaks_el.text:
            prec = int(peaks_el.get("precision", "32"))
            compr = peaks_el.get("compressionType", "none").lower()
            bo = peaks_el.get("byteOrder", "network").lower()
            flat = _decode_binary(peaks_el.text, "zlib" in compr, prec, bo == "network")
            mz_arr = flat[0::2]
            int_arr = flat[1::2]
        if lvl == 1:
            if mz_arr and int_arr:
                mz_np = np.asarray(mz_arr, dtype=np.float64)
                in_np = np.asarray(int_arr, dtype=np.float64)
                order = np.argsort(mz_np)
                ms1.append({"rt_min": rt_min, "mz": mz_np[order], "inten": in_np[order]})
        elif lvl == 2:
            prec_mz = 0.0
            prec_el = elem.find("precursorMz")
            if prec_el is not None:
                try:
                    prec_mz = float(prec_el.text or 0)
                except Exception:
                    pass
            pairs = sorted(zip(mz_arr, int_arr), key=lambda x: x[0]) if mz_arr and int_arr else []
            ms2.append({"rt_min": rt_min, "prec_mz": prec_mz, "pairs": pairs})
        elem.clear()
    ms1.sort(key=lambda s: s["rt_min"])
    ms2.sort(key=lambda s: s["rt_min"])
    return ms1, ms2


# ═════════════════════════════════════════════════════════════════════════════
# MGF READER  — MGF spectra are ALREADY peak-picked (one precursor per BEGIN/END
# IONS block, with its RT + fragments). There is no MS1 chromatographic axis, so
# alignment treats each spectrum directly as one detected peak — no ROI building,
# no peak detection. This is the fastest path by far (parse-only).
# ═════════════════════════════════════════════════════════════════════════════
def read_mgf_spectra(path, cb=None):
    """Parse an MGF file. Returns a list of
    {'rt_min','prec_mz','prec_int','frag_text'}.

    SPEED: fragment m/z-intensity pairs are NOT parsed here — they are kept as
    the raw text block (`frag_text`) and parsed lazily via `_get_pairs` only
    for the handful of spectra actually chosen for MS2 output (one per feature).
    A typical MGF has ~240 fragments/spectrum, so eager parsing would convert
    millions of floats up front (the dominant cost); lazy parsing skips ~99%
    of them. Header parsing stops at the first fragment line, so we never even
    touch the fragment lines during the main pass."""
    with open(path, "r", errors="replace") as fh:
        text = fh.read()

    spectra = []
    blocks = text.split("BEGIN IONS")
    for blk in blocks[1:]:
        e = blk.find("END IONS")
        seg = blk[:e] if e != -1 else blk
        lines = seg.split("\n")
        prec_mz = 0.0
        prec_int = 0.0
        rt = 0.0
        frag_idx = None
        for i, raw in enumerate(lines):
            line = raw.strip()
            if not line:
                continue
            if "=" in line:
                key, _, val = line.partition("=")
                key = key.strip().upper()
                val = val.strip()
                if key == "PEPMASS":
                    parts = val.split()
                    try:
                        prec_mz = float(parts[0])
                    except (ValueError, IndexError):
                        pass
                    if len(parts) > 1:
                        try:
                            prec_int = float(parts[1])
                        except ValueError:
                            pass
                elif key == "RTINSECONDS":
                    try:
                        rt = float(val) / 60.0
                    except ValueError:
                        pass
                elif key == "RTINMINUTES" and rt == 0.0:
                    try:
                        rt = float(val)
                    except ValueError:
                        pass
            else:
                # First non-key line = start of the fragment block (MGF puts
                # all headers before the peak list). Stop header parsing here.
                frag_idx = i
                break
        if prec_mz <= 0:
            continue
        frag_text = "\n".join(lines[frag_idx:]) if frag_idx is not None else ""
        spectra.append({"rt_min": rt, "prec_mz": prec_mz,
                        "prec_int": prec_int, "frag_text": frag_text})
    if cb:
        cb(f"  {len(spectra):,} MGF spectra parsed.")
    return spectra


def _parse_frag_text(frag_text):
    """Lazily parse an MGF fragment block into sorted (mz, intensity) pairs."""
    pairs = []
    for ln in frag_text.split("\n"):
        parts = ln.split()
        if len(parts) >= 2:
            try:
                pairs.append((float(parts[0]), float(parts[1])))
            except ValueError:
                pass
    pairs.sort(key=lambda x: x[0])
    return pairs


def _get_pairs(spectrum):
    """Return fragment (mz, intensity) pairs from either an already-parsed
    scan ('pairs', from mzML/mzXML) or a lazy MGF spectrum ('frag_text')."""
    if "pairs" in spectrum:
        return spectrum["pairs"]
    return _parse_frag_text(spectrum.get("frag_text", ""))


def peaks_from_mgf(spectra):
    """Turn MGF spectra directly into peak dicts (one peak per spectrum).
    Height is the precursor intensity if present, else the base fragment
    intensity, else 1.0 (so intensity-weighted grouping still works)."""
    peaks = []
    for sp in spectra:
        mz = sp["prec_mz"]
        rt = sp["rt_min"]
        height = sp["prec_int"]
        if height <= 0:
            height = max((b for _, b in _get_pairs(sp)), default=1.0)
        peaks.append({
            "mz": mz, "mz_min": mz, "mz_max": mz,
            "rt": rt, "rt_min": rt, "rt_max": rt,
            "height": float(height), "area": float(height),
            "snr": float("inf"),
        })
    return peaks


# ═════════════════════════════════════════════════════════════════════════════
# RAW CONVERSION  — ThermoRawFileParser only (msConvert is not assumed to be
# installed on target machines). Runs via Popen with periodic heartbeat log
# lines and a hard timeout, so a slow/large file never looks like a frozen UI.
# ═════════════════════════════════════════════════════════════════════════════
TRFP_DIR = Path.home() / "VeroMass_Aligner" / "ThermoRawFileParser"
RAW_CONVERT_TIMEOUT_SEC = 1800   # hard cap; fails with a clear error rather than hanging forever
HEARTBEAT_SEC = 10


def find_trfp(custom=""):
    candidates = (
        ([Path(custom)] if custom else [])
        + list(TRFP_DIR.rglob("ThermoRawFileParser.exe"))
        + [Path("ThermoRawFileParser.exe")]
    )
    for p in candidates:
        if Path(p).exists():
            return str(p)
    return None


def get_trfp_url():
    try:
        r = req.get(
            "https://api.github.com/repos/compomics/ThermoRawFileParser/releases/latest",
            timeout=10, headers={"Accept": "application/vnd.github.v3+json"},
        )
        assets = r.json().get("assets", [])
        for a in assets:
            nm = a["name"].lower()
            if nm.endswith(".zip") and any(w in nm for w in ("win", "windows")):
                return a["browser_download_url"]
        for a in assets:
            nm = a["name"].lower()
            if nm.endswith(".zip") and not any(w in nm for w in ("linux", "mac", "osx", "darwin", "arm")):
                return a["browser_download_url"]
        for a in assets:
            if a["name"].lower().endswith(".zip"):
                return a["browser_download_url"]
    except Exception:
        pass
    return ("https://github.com/compomics/ThermoRawFileParser/releases/"
            "download/v1.4.5/ThermoRawFileParser1.4.5.zip")


def download_trfp(cb):
    TRFP_DIR.mkdir(parents=True, exist_ok=True)
    url = get_trfp_url()
    fname = url.split("/")[-1]
    cb(f"Downloading {fname} …")
    r = req.get(url, stream=True, timeout=300)
    r.raise_for_status()
    zp = TRFP_DIR / "trfp.zip"
    with open(zp, "wb") as f:
        for chunk in r.iter_content(512 * 1024):
            f.write(chunk)
    cb("Extracting…")
    with zipfile.ZipFile(zp, "r") as z:
        z.extractall(TRFP_DIR)
    zp.unlink(missing_ok=True)
    exe = find_trfp()
    if exe:
        cb("ThermoRawFileParser ready.")
        return exe
    found = [f.name for f in TRFP_DIR.rglob("*") if f.is_file()]
    raise RuntimeError(f"ThermoRawFileParser.exe not found. Extracted: {found[:8]}")


def _run_with_heartbeat(cmd, cb, timeout_sec):
    """Run cmd via Popen, emitting a heartbeat log line every HEARTBEAT_SEC so
    a long-running conversion never looks frozen, and hard-killing the process
    if it exceeds timeout_sec instead of blocking forever."""
    # On Windows, run the converter in its OWN process group and with no console
    # window. Without CREATE_NEW_PROCESS_GROUP the child shares the parent's
    # console group, so a single Ctrl+C / console break kills ALL in-flight
    # converters at once with exit 3221225786 (0xC000013A, STATUS_CONTROL_C_EXIT)
    # — the exact cascade that left truncated .mzML stubs poisoning the cache.
    # CREATE_NO_WINDOW also stops each concurrent conversion from flashing its
    # own console window when the app is launched via pythonw/.vbs.
    creationflags = 0
    if sys.platform == "win32":
        creationflags = subprocess.CREATE_NEW_PROCESS_GROUP | subprocess.CREATE_NO_WINDOW
    proc = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                            text=True, creationflags=creationflags)
    start = time.time()
    last_beat = start
    while True:
        ret = proc.poll()
        if ret is not None:
            out = proc.stdout.read() if proc.stdout else ""
            return ret, out
        now = time.time()
        if now - start > timeout_sec:
            proc.kill()
            raise RuntimeError(f"Conversion exceeded {timeout_sec}s timeout — killed.")
        if now - last_beat > HEARTBEAT_SEC:
            elapsed = int(now - start)
            cb(f"  …still converting ({elapsed}s elapsed)")
            last_beat = now
        time.sleep(0.5)


def _mzml_complete(path):
    """Cheap integrity check: a well-formed indexed mzML ends with
    </indexedmzML> (or at least </mzML>). A conversion killed mid-write leaves a
    stub that ends inside a <spectrum>, so a tail check reliably rejects partials
    without parsing the whole file. This is what stops a truncated cache entry
    from being silently reused (the 'no element found: line N' failures)."""
    try:
        if os.path.getsize(path) < 1024:
            return False
        with open(path, "rb") as fh:
            fh.seek(max(0, os.path.getsize(path) - 512))
            tail = fh.read()
        return b"</indexedmzML>" in tail or b"</mzML>" in tail
    except OSError:
        return False


def convert_raw(raw_path, out_dir, cb, custom_trfp=""):
    raw_path = Path(raw_path)
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    expected = out_dir / (raw_path.stem + ".mzML")

    # Reuse a cache entry ONLY if it is actually complete. A partial stub left by
    # an interrupted earlier run must never be trusted just because it exists.
    if expected.exists():
        if _mzml_complete(expected):
            cb(f"  Using cached mzML: {expected.name}")
            return str(expected)
        cb(f"  Discarding incomplete cached mzML: {expected.name}")
        try:
            expected.unlink()
        except OSError:
            pass

    trfp = find_trfp(custom_trfp)
    if not trfp:
        trfp = download_trfp(cb)
    cb(f"  Converting via ThermoRawFileParser: {raw_path.name}…")

    # Convert to a temp file in the same directory, validate, then atomically
    # rename into place (os.replace). A crash or interrupt can then only leave a
    # *.part.mzML partial (removed below) — never a poisoned `expected` cache
    # entry. We also address explicitly to `-b=<tmp>` and return `expected`
    # itself rather than globbing the shared cache dir and taking the
    # alphabetically-last .mzML (which, with many files in one cache dir, could
    # hand back a DIFFERENT sample's data).
    tmp = out_dir / (raw_path.stem + ".part.mzML")
    try:
        tmp.unlink()
    except OSError:
        pass
    cmd = [trfp, f"-i={raw_path}", f"-b={tmp}", "-f=2", "-m=2"]
    ret, out = _run_with_heartbeat(cmd, cb, RAW_CONVERT_TIMEOUT_SEC)
    if ret != 0 or not tmp.exists() or not _mzml_complete(tmp):
        try:
            tmp.unlink()
        except OSError:
            pass
        raise RuntimeError(f"Conversion failed (exit {ret}):\n{out[-300:]}")
    os.replace(str(tmp), str(expected))
    return str(expected)


# ═════════════════════════════════════════════════════════════════════════════
# PEAK DETECTION  (ROI building + multi-scale peak picking)
# ═════════════════════════════════════════════════════════════════════════════
_trapz = getattr(np, "trapezoid", None) or np.trapz


class PeakDetectionParams:
    def __init__(self, ppm=25, peak_min_sec=1, peak_max_sec=60,
                 snr_thresh=3.0, noise=1000, prefilter_k=4, prefilter_i=100,
                 max_gap_scans=3):
        self.ppm = ppm
        self.peak_min_sec = peak_min_sec
        self.peak_max_sec = peak_max_sec
        self.snr_thresh = snr_thresh
        self.noise = noise
        self.prefilter_k = prefilter_k
        self.prefilter_i = prefilter_i
        # Max consecutive-scan gap tolerated inside one ROI before it is split
        # into separate elution windows. Without this an ion's real peak gets
        # merged with all its scattered background across the whole run.
        self.max_gap_scans = max_gap_scans


def _build_rois(scans, ppm, prefilter_k, prefilter_i, cb=None, max_gap_scans=3):
    """
    Build regions-of-interest (per-ion extracted chromatograms) in a single
    VECTORIZED pass, returning (rois, rts).

    Each ROI is {'scan': ndarray of scan indices, 'mz': ndarray, 'inten':
    ndarray}, and `rts` maps scan index -> retention time (minutes).

    ALGORITHM (this is the performance breakthrough): instead of the classic
    incremental centWave loop — which carries a growing set of "active" ROIs
    and, for every m/z point in every scan, scans that active set (O(scans ×
    active_ROIs), which ballooned to ~70k active ROIs and ~90s/file) — we:
      1. Flatten every MS1 point (m/z, intensity, scan index) into three flat
         arrays.
      2. Drop points below the intensity prefilter globally (removes the vast
         majority of noise points in one vectorized mask).
      3. Sort once by m/z, then split into ROIs wherever the m/z gap between
         consecutive points exceeds the local ppm tolerance (np.diff + np.where
         — fully vectorized; ion points cluster tightly in m/z so splits land
         in empty m/z space).
      4. Collapse each ROI onto its scan axis (max intensity per scan) with
         np.maximum.at / np.add.at.
    This is O(N log N) in the total number of points — no Python loop over an
    active-ROI set — and runs in ~1s for a 2,300-scan file that previously
    took ~90s. Memory is O(total_points) (flat arrays), never the dense
    O(n_bins × n_scans) matrix that OOM-crashed an earlier implementation.
    """
    n_scans = len(scans)
    rts = np.array([s["rt_min"] for s in scans], dtype=np.float64)
    total = int(sum(s["mz"].size for s in scans))
    if total == 0:
        return [], rts

    mz_all = np.empty(total, dtype=np.float64)
    int_all = np.empty(total, dtype=np.float64)
    scan_all = np.empty(total, dtype=np.int64)
    pos = 0
    for si, s in enumerate(scans):
        k = s["mz"].size
        if not k:
            continue
        mz_all[pos:pos + k] = s["mz"]
        int_all[pos:pos + k] = s["inten"]
        scan_all[pos:pos + k] = si
        pos += k

    keep = int_all >= prefilter_i
    mz_all = mz_all[keep]
    int_all = int_all[keep]
    scan_all = scan_all[keep]
    if mz_all.size == 0:
        return [], rts

    order = np.argsort(mz_all, kind="quicksort")
    mz_s = mz_all[order]
    int_s = int_all[order]
    scan_s = scan_all[order]

    tol = mz_s[:-1] * ppm * 1e-6
    gaps = np.where(np.diff(mz_s) > tol)[0] + 1
    starts = np.concatenate(([0], gaps))
    ends = np.concatenate((gaps, [mz_s.size]))

    rois = []
    for a, b in zip(starts, ends):
        if b - a < prefilter_k:
            continue
        sc = scan_s[a:b]
        it = int_s[a:b]
        mzz = mz_s[a:b]
        uscan, inv = np.unique(sc, return_inverse=True)
        if uscan.size < prefilter_k:
            continue
        eic = np.zeros(uscan.size)
        np.maximum.at(eic, inv, it)
        mzsum = np.zeros(uscan.size)
        wsum = np.zeros(uscan.size)
        np.add.at(mzsum, inv, mzz * it)
        np.add.at(wsum, inv, it)
        mz_eic = mzsum / np.maximum(wsum, 1e-9)

        # ── Split into time-contiguous elution windows ───────────────────────
        # An m/z bucket spans the WHOLE run: a compound's real 10-second peak
        # plus every scattered background occurrence of that same m/z. Treating
        # that as one ROI makes the trace's own baseline statistics meaningless
        # (measured: median 5 separate elution segments per bucket, worst 248),
        # which both suppresses real peaks and invents noise ones. Split
        # wherever the scan index jumps more than max_gap_scans.
        brk = np.where(np.diff(uscan) > max_gap_scans)[0] + 1
        for seg_scan, seg_mz, seg_int in zip(np.split(uscan, brk),
                                             np.split(mz_eic, brk),
                                             np.split(eic, brk)):
            if seg_scan.size >= prefilter_k:
                rois.append({"scan": seg_scan, "mz": seg_mz, "inten": seg_int})
    return rois, rts


def _detect_peaks_in_roi(roi, rts, params):
    """
    Gaussian-smoothed local-maxima peak picking along a ROI's native
    RT-intensity trace (no fixed-grid upsampling — the scan axis IS the grid,
    which is both faster and more faithful than interpolating). Returns peak
    dicts with RT/mz bounds, height, area, S/N.
    """
    uscan = roi["scan"]
    inten = roi["inten"]
    mz = roi["mz"]
    rt = rts[uscan]
    if rt.size < 3:
        return []

    dt = np.median(np.diff(rt))
    if dt <= 0:
        return []

    smoothed = gaussian_filter1d(inten, sigma=1.0)

    # Baseline + robust noise sigma estimated from the LOW half of this elution
    # window. The previous formula (max(noise, median(inten)*0.5)) derived the
    # "noise floor" from the trace's own median — so a well-sampled, genuinely
    # eluting compound defined its own noise and was rejected at S/N ~2. That
    # cost ~76% of unambiguous peaks. Baseline = p25; sigma = MAD of the lower
    # half, scaled to a normal-equivalent standard deviation.
    # One sort serves both statistics — np.percentile on many tiny arrays is
    # dominated by call overhead, and this runs once per ROI across ~36k ROIs.
    srt = np.sort(inten)
    n_pts = srt.size
    baseline = float(srt[int(0.25 * (n_pts - 1))])
    lower = srt[:max(n_pts // 2, 1)]
    if lower.size > 2:
        med_l = lower[lower.size // 2]
        sigma = float(np.median(np.abs(lower - med_l))) * 1.4826
    else:
        sigma = 0.0
    sigma = max(sigma, 1.0)

    min_dist_pts = max(int((params.peak_min_sec / 60.0) / dt), 1)
    # NOTE: deliberately no `prominence` filter here. Topographic prominence is
    # measured against neighbouring maxima, not against the baseline, so it
    # rejected genuine peaks purely for eluting close to another peak — it was
    # the single largest source of missed features (29% of all real peaks),
    # while being redundant with the S/N test below, which is the
    # scientifically meaningful criterion (height above baseline, in sigmas).
    peaks_idx, props = find_peaks(
        smoothed,
        height=max(params.noise, baseline),
        distance=min_dist_pts,
        width=1,
    )
    if peaks_idx.size == 0:
        return []

    results = []
    widths_res = props["widths"]
    left_ips = props["left_ips"]
    right_ips = props["right_ips"]
    n = rt.size

    for k, pk in enumerate(peaks_idx):
        width_sec = float(widths_res[k]) * dt * 60.0
        if width_sec < params.peak_min_sec or width_sec > params.peak_max_sec:
            continue
        li = int(np.clip(np.floor(left_ips[k]), 0, n - 1))
        ri = int(np.clip(np.ceil(right_ips[k]), 0, n - 1))
        if ri <= li:
            ri = min(li + 1, n - 1)

        # Report the TRUE MEASURED apex, never the smoothed value. Smoothing is
        # a detection aid (it locates the peak); reporting smoothed intensity
        # systematically understates every peak — measured on real data: median
        # 17% low, worst decile 29% low, total ion intensity only 85% of true.
        # The apex is taken as the max raw intensity inside the peak bounds.
        seg = inten[li:ri + 1]
        apex_off = int(np.argmax(seg))
        apex_i = li + apex_off
        height = float(seg[apex_off])

        snr = (height - baseline) / sigma
        if snr < params.snr_thresh:
            continue

        area = float(_trapz(seg, rt[li:ri + 1]) * 60.0)
        mz_seg = mz[li:ri + 1]
        mz_apex = float(np.average(mz_seg, weights=seg)) if seg.sum() > 0 else float(mz[apex_i])

        results.append({
            "mz": mz_apex, "mz_min": float(mz_seg.min()), "mz_max": float(mz_seg.max()),
            "rt": float(rt[apex_i]), "rt_min": float(rt[li]), "rt_max": float(rt[ri]),
            "height": height, "area": area, "snr": snr,
        })
    return results


def detect_chrom_peaks(scans, params, cb=None):
    """Full per-file peak detection: vectorized ROI building + peak picking."""
    if cb:
        cb(f"  Building ROIs from {len(scans)} MS1 scans…")
    rois, rts = _build_rois(scans, params.ppm, params.prefilter_k, params.prefilter_i,
                             cb=cb, max_gap_scans=params.max_gap_scans)
    if cb:
        cb(f"  {len(rois)} candidate ROI(s); detecting peaks…")
    peaks = []
    for roi in rois:
        peaks.extend(_detect_peaks_in_roi(roi, rts, params))
    if cb:
        cb(f"  {len(peaks)} chromatographic peak(s) detected.")
    return peaks


# ═════════════════════════════════════════════════════════════════════════════
# CORRESPONDENCE  (cross-sample density-based grouping)
# ═════════════════════════════════════════════════════════════════════════════
class GroupingParams:
    def __init__(self, mz_ppm=15, mz_abs=0.005, rt_bw_sec=15, min_frac_samples=0.5,
                 rt_capture_mult=0.1):
        self.mz_ppm = mz_ppm
        self.mz_abs = mz_abs
        self.rt_bw_sec = rt_bw_sec
        self.min_frac_samples = min_frac_samples
        # Capture radius for group_peaks_density, as a multiple of rt_bw_sec.
        # Verified against the real xcms reference report (60,512 features,
        # see project notes): 0.1 gave the best result on EVERY axis at once
        # (73.5% match vs 73.1% for the old scheme, RT bias +0.0001min vs
        # +0.065min, tightest RT std, and zero structurally-illegal
        # `size>2` clusters in a 2-sample run vs 32,063 before). Wider values
        # (0.3-2.0) all measured WORSE on match rate despite also fixing RT
        # bias — capturing too widely folds real, distinct, nearby
        # compounds into one density maximum. Do not widen this without
        # re-running that comparison.
        self.rt_capture_mult = rt_capture_mult


def _mz_tol(mz, params):
    return max(params.mz_abs, mz * params.mz_ppm * 1e-6)


def group_peaks(all_peaks, n_samples, params, cb=None):
    """
    all_peaks: list of dicts each with an added 'sample' key (index of source file).
    Groups peaks into cross-sample features: m/z-tolerance bucketing, then
    RT-density clustering within each bucket. Returns list of feature dicts.

    NOTE: an attempt was made (and reverted) to replace this with pure
    gap-based splitting (sort + split wherever the consecutive gap exceeds
    tolerance), on the theory that it would be less "order-dependent" than
    the seed-window scheme below. It is NOT a safe substitute: gap-based
    splitting is mathematically single-linkage/chain clustering — exactly
    the failure mode the seed-window fix below exists to prevent — and
    measured WORSE against the real xcms reference (54.3% vs 73.1% match,
    with `size>2` features reappearing in a 2-sample run where that's
    structurally impossible for a correct clustering). The genuine
    order-dependent "bucket stealing" bug (a real cross-sample match losing
    a member to an earlier unrelated seed's greedy claim) is still open and
    needs a real best-match/bipartite-style rework, not a gap-based swap —
    do not attempt that swap again without re-measuring against the full
    xcms reference file first.
    """
    if not all_peaks:
        return []

    peaks_sorted = sorted(all_peaks, key=lambda p: p["mz"])
    n = len(peaks_sorted)
    used = [False] * n
    features = []
    bw_min = params.rt_bw_sec / 60.0

    i = 0
    while i < n:
        if used[i]:
            i += 1
            continue
        # FIXED WINDOW FROM THE SEED, NOT CHAIN-EXPANDING. The previous version
        # grew hi_mz every time a new point extended it ("hi_mz = max(hi_mz,
        # peaks_sorted[j]['mz'] + tol(j))"), which is single-linkage/chain
        # clustering: it only bounds the gap between CONSECUTIVE points, not
        # total cluster width. With ~100k+ peaks in a run, a chain of
        # closely-spaced points can link a cluster far wider than the intended
        # ppm tolerance, silently merging unrelated species. Confirmed in
        # production: a feature reported Size=7 from a 2-sample alignment
        # (should be at most 2), and its inflated cluster corrupted the RT
        # anchor set used for RT correction, which then applied a spurious
        # +0.24 to +0.35 min shift to EVERY peak in both samples (verified
        # against a real xcms comparison — see project notes). The window is
        # now fixed to the seed's own tolerance and never grows.
        seed = peaks_sorted[i]
        tol = _mz_tol(seed["mz"], params)
        hi_mz = seed["mz"] + tol
        bucket_idx = [i]
        j = i + 1
        while j < n and peaks_sorted[j]["mz"] <= hi_mz:
            if not used[j]:
                bucket_idx.append(j)
            j += 1

        bucket_idx.sort(key=lambda k: peaks_sorted[k]["rt"])
        # Same fix for the RT sub-clustering: bound each sub-cluster's width
        # to bw_min from its FIRST member (cur[0]), not from the last-added
        # member (cur[-1]) — the latter is the same chain-clustering flaw and
        # let RT sub-clusters drift arbitrarily wide via a chain of
        # consecutively-close points.
        sub_clusters = []
        cur = [bucket_idx[0]]
        for k in bucket_idx[1:]:
            if peaks_sorted[k]["rt"] - peaks_sorted[cur[0]]["rt"] <= bw_min:
                cur.append(k)
            else:
                sub_clusters.append(cur)
                cur = [k]
        sub_clusters.append(cur)

        for cluster_idx in sub_clusters:
            cluster = [peaks_sorted[k] for k in cluster_idx]
            n_samples_in_cluster = len(set(p["sample"] for p in cluster))
            if n_samples_in_cluster / max(n_samples, 1) < params.min_frac_samples:
                continue
            for k in cluster_idx:
                used[k] = True
            mzs = np.array([p["mz"] for p in cluster])
            rts = np.array([p["rt"] for p in cluster])
            heights = np.array([p["height"] for p in cluster])
            mz_mins = np.array([p["mz_min"] for p in cluster])
            mz_maxs = np.array([p["mz_max"] for p in cluster])
            rt_mins = np.array([p["rt_min"] for p in cluster])
            rt_maxs = np.array([p["rt_max"] for p in cluster])

            shape_distance = float(np.std(rts) / max(np.mean(rts), 1e-9))
            best_peak = max(cluster, key=lambda p: p["height"])

            features.append({
                "peaks": cluster,
                "size": len(cluster),
                "mz": float(np.average(mzs, weights=heights)),
                "mz_min": float(mz_mins.min()),
                "mz_max": float(mz_maxs.max()),
                "rt": float(np.average(rts, weights=heights)),
                "rt_min": float(rt_mins.min()),
                "rt_max": float(rt_maxs.max()),
                "base_peak": float(heights.max()),
                "shape_distance": shape_distance,
                "rep_sample": best_peak["sample"],
                "rep_rt": best_peak["rt"],
                "rep_mz": best_peak["mz"],
            })
        i = j if j > i else i + 1

    return features


class PeakStore:
    """Columnar (numpy) peak storage — replaces per-peak Python dicts.

    A peak dict was ~500-900 B; these parallel columns are 70 B/peak (measured),
    so a 300-file run (33.6M peaks) drops from ~17 GB (OOM on 16 GB) to ~2.4 GB.
    Only `height` is float32 — it is the ONLY field verified float32-LOSSLESS
    across the real 1.57M-peak run; every other field needs float64 to keep the
    grouping/RT/export output BYTE-IDENTICAL to the old dict pipeline (verified by
    feature-set fingerprint on 2/4/8/14-sample subsets). Compute sites upcast
    height to float64, which is exact, so np.average / KDE bits are unchanged.

    Features reference peaks by GLOBAL index (feat["peak_idx"]) rather than
    holding dict references, so the peaks live ONLY here and the memory is
    actually reclaimed. rt/rt_min/rt_max are mutated in place by
    apply_rt_correction (per-peak scalar arithmetic, see there)."""
    __slots__ = ("mz", "mz_min", "mz_max", "rt", "rt_min", "rt_max",
                 "height", "area", "snr", "sample", "n")
    _F64 = ("mz", "mz_min", "mz_max", "rt", "rt_min", "rt_max", "area", "snr")

    def __init__(self, n):
        self.n = n
        for f in self._F64:
            setattr(self, f, np.empty(n, dtype=np.float64))
        self.height = np.empty(n, dtype=np.float32)
        self.sample = np.empty(n, dtype=np.int16)

    @classmethod
    def from_dicts(cls, peaks):
        """Build a store (or a per-file chunk) from a list of peak dicts, each
        carrying a 'sample' index. The dicts can be freed by the caller right
        after — this is how run_alignment keeps only one file's dicts alive at a
        time while the columnar store grows."""
        n = len(peaks)
        s = cls(n)
        if n:
            s.mz[:]     = [p["mz"] for p in peaks]
            s.mz_min[:] = [p["mz_min"] for p in peaks]
            s.mz_max[:] = [p["mz_max"] for p in peaks]
            s.rt[:]     = [p["rt"] for p in peaks]
            s.rt_min[:] = [p["rt_min"] for p in peaks]
            s.rt_max[:] = [p["rt_max"] for p in peaks]
            s.area[:]   = [p["area"] for p in peaks]
            s.snr[:]    = [p["snr"] for p in peaks]
            s.height[:] = [p["height"] for p in peaks]   # float32 store (lossless)
            s.sample[:] = [p["sample"] for p in peaks]
        return s

    @classmethod
    def concat(cls, chunks):
        """Concatenate per-file chunks into one store, preserving order (so the
        stable m/z argsort below reproduces the old sorted(all_peaks) tie-break
        of file-0 peaks before file-1, etc.)."""
        chunks = [c for c in chunks if c.n]
        if not chunks:
            return cls(0)
        s = cls(sum(c.n for c in chunks))
        for f in cls._F64:
            np.concatenate([getattr(c, f) for c in chunks], out=getattr(s, f))
        np.concatenate([c.height for c in chunks], out=s.height)
        np.concatenate([c.sample for c in chunks], out=s.sample)
        return s

    def nbytes(self):
        tot = self.height.nbytes + self.sample.nbytes
        for f in self._F64:
            tot += getattr(self, f).nbytes
        return tot


def _feature_from_cluster(store, idxs):
    """Build a feature from a cluster given as GLOBAL peak indices into `store`.
    Stores 'peak_idx' (the index array) instead of a list of dicts. Element order
    in the gathered arrays == idxs order == the winners order from
    _density_cluster_bin, so the floating-point summation in np.average and the
    first-max tie-break in argmax match the old dict pipeline exactly."""
    mzs      = store.mz[idxs]
    rts      = store.rt[idxs]
    heights  = store.height[idxs].astype(np.float64)   # exact (float32-lossless)
    mz_mins  = store.mz_min[idxs]
    mz_maxs  = store.mz_max[idxs]
    rt_mins  = store.rt_min[idxs]
    rt_maxs  = store.rt_max[idxs]
    shape_distance = float(np.std(rts) / max(np.mean(rts), 1e-9))
    best_local = int(np.argmax(heights))               # first-max == max(key=height)
    best_gi = int(idxs[best_local])
    return {
        "peak_idx": idxs,
        "size": int(idxs.size),
        "mz": float(np.average(mzs, weights=heights)),
        "mz_min": float(mz_mins.min()),
        "mz_max": float(mz_maxs.max()),
        "rt": float(np.average(rts, weights=heights)),
        "rt_min": float(rt_mins.min()),
        "rt_max": float(rt_maxs.max()),
        "base_peak": float(heights.max()),
        "shape_distance": shape_distance,
        "rep_sample": int(store.sample[best_gi]),
        "rep_rt": float(store.rt[best_gi]),
        "rep_mz": float(store.mz[best_gi]),
    }


def _density_cluster_bin(store, gidx, n_samples, bw_min, min_frac_samples, capture_mult=1.0):
    """
    Density-based RT correspondence within ONE coarse m/z slice — the same
    family of algorithm xcms's own PeakDensityParam uses, chosen specifically
    because it has no "seed"/"first mover": every point's cluster assignment
    falls out of the shape of a kernel density estimate built from ALL points
    in the bin at once, so processing order cannot change the result (unlike
    the seed-window scheme in `group_peaks`, which is still known to let an
    unrelated peak be greedily claimed by the wrong seed — see `group_peaks`'s
    docstring for the measured evidence). It also resists chain-merging by
    construction: a real gap between two compounds shows up as a genuine dip
    between two density maxima, while a few incidental noise points bridging
    them get smoothed away by the Gaussian bandwidth rather than read as one
    continuous single-linkage chain.

    Algorithm (iterative density-maxima peeling):
      1. Build a Gaussian-kernel density estimate over RT, weighted by peak
         intensity, bandwidth = bw_min (same knob as the existing rt_bw_sec).
      2. Take the global maximum of that density curve as the next feature's
         reference RT.
      3. Capture every remaining peak within `capture_mult * bw_min` of it.
      4. Enforce one-peak-per-sample: if a sample contributes more than one
         captured peak, keep only its highest-intensity one and return the
         rest to the pool — they may belong to a distinct, nearby compound
         (e.g. an isomer) that deserves its own density maximum later.
      5. Remove every captured (kept or returned) peak from the pool,
         re-estimate density on what's left, and repeat until the pool is
         empty or the remaining weight is negligible.
      6. Apply the existing min_frac_samples filter per resulting cluster.
    """
    if gidx.size == 0:
        return []
    # Peaks held as fixed numpy columns with a boolean alive-mask instead of
    # rebuilding Python lists from dicts every iteration; the KDE density is
    # maintained INCREMENTALLY (removed peaks' kernels are subtracted) and only
    # rebuilt when the alive RT extremes change (i.e. when the grid domain
    # actually shifts). This flattens the old O(features x peaks) per-bin cost to
    # near-linear (measured peaks^0.81 on a real 1.57M-peak run) while producing
    # byte-identical clusters — verified by fingerprint against the original on
    # 2/4/8/14-sample subsets. Do NOT reintroduce the per-iteration full rebuild.
    # gidx = GLOBAL peak indices for this m/z bin, in m/z-sorted order (== the old
    # bin order). Gather compute columns as float64; upcasting the float32 height
    # store is exact (lossless-verified), so these arrays are bit-identical to the
    # old np.fromiter over dicts.
    M = gidx.size
    rt_arr = store.rt[gidx].astype(np.float64, copy=True)
    ht_arr = store.height[gidx].astype(np.float64, copy=True)
    sm_arr = store.sample[gidx].astype(np.int64, copy=True)
    alive = np.ones(M, dtype=bool)

    grid_step = bw_min / 4.0
    capture_r = capture_mult * bw_min
    clusters = []

    grid = dens = None
    cur_lo = cur_hi = None

    while True:
        idx = np.nonzero(alive)[0]           # original order preserved
        if idx.size == 0:
            break
        rts = rt_arr[idx]
        weights = ht_arr[idx]
        if weights.sum() <= 0:
            break
        amin = float(rts.min()); amax = float(rts.max())
        # Rebuild the grid + full KDE only when the alive RT range moved; the
        # exact same grid is required for byte-identical argmax/capture results.
        if grid is None or amin != cur_lo or amax != cur_hi:
            lo, hi = amin - bw_min, amax + bw_min
            grid = np.arange(lo, hi + grid_step, grid_step)
            diff = (grid[:, None] - rts[None, :]) / bw_min
            dens = (weights[None, :] * np.exp(-0.5 * diff * diff)).sum(axis=1)
            cur_lo, cur_hi = amin, amax
        peak_i = int(np.argmax(dens))
        if dens[peak_i] <= 0:
            break
        center_rt = grid[peak_i]

        cap_local = np.nonzero(np.abs(rts - center_rt) <= capture_r)[0]
        if cap_local.size == 0:
            # numerical edge case — drop the single nearest point to guarantee progress
            cap_local = np.array([int(np.argmin(np.abs(rts - center_rt)))])

        # one peak per sample: highest height wins, first-seen wins ties (exactly
        # the dict-insertion semantics of the original scan-order loop).
        by_sample = {}
        for k in cap_local:
            gi = int(idx[k])
            s = int(sm_arr[gi])
            if s not in by_sample or ht_arr[gi] > ht_arr[by_sample[s]]:
                by_sample[s] = gi
        winners = list(by_sample.values())

        if len(by_sample) / max(n_samples, 1) >= min_frac_samples:
            # winners are bin-local indices; map to GLOBAL store indices via gidx.
            clusters.append(gidx[np.array(winners, dtype=np.int64)])

        # Remove ONLY the kept per-sample winners (same as the original). The
        # same-sample NON-winners inside the capture radius are deliberately left
        # alive so a later density maximum can claim them — a dropped real peak is
        # unrecoverable downstream (this is the v1.8.1 silent-drop fix: do NOT
        # remove all captured points). Subtract the winners' kernels from the
        # maintained density; if a winner was an RT extreme, the next iteration
        # detects the changed range and rebuilds exactly. Winners are always
        # non-empty, so `alive` strictly shrinks and termination holds.
        wr = rt_arr[winners]; ww = ht_arr[winners]
        d2 = (grid[:, None] - wr[None, :]) / bw_min
        dens = dens - (ww[None, :] * np.exp(-0.5 * d2 * d2)).sum(axis=1)
        alive[winners] = False

    return clusters


def group_peaks_density(store, n_samples, params, cb=None):
    """
    Cross-sample correspondence via density-based clustering (see
    `_density_cluster_bin`). Coarse m/z partitioning still uses gap-based
    splitting (safe here — proven at scale in `_build_rois` — because the
    fine-grained, order-sensitive separation of distinct compounds happens
    in the RT-density step below, not in this m/z pass, which only needs to
    keep obviously-unrelated masses out of the same density estimate).
    Verified against the real xcms reference report (60,512 features from a
    real study): matches the old seed-window scheme's raw match rate
    (73.5% vs 73.1%) while collapsing the systemic RT bias from +0.065min to
    +0.0001min and eliminating all structurally-illegal `size>2` clusters
    in a 2-sample test (32,063 -> 0). See GroupingParams.rt_capture_mult for
    the capture-radius tuning history.
    """
    n = store.n
    if n == 0:
        return []

    bw_min = params.rt_bw_sec / 60.0
    # stable argsort reproduces the old sorted(all_peaks, key=mz): equal-m/z peaks
    # keep their original (file-0-before-file-1) order, so bin edges and the
    # one-peak-per-sample tie-breaks below are byte-identical to the dict version.
    order = np.argsort(store.mz, kind="stable")
    mzs_all = store.mz[order]

    tol = np.maximum(mzs_all[:-1] * params.mz_ppm * 1e-6, params.mz_abs)
    mz_gaps = np.where(np.diff(mzs_all) > tol)[0] + 1
    mz_starts = np.concatenate(([0], mz_gaps))
    mz_ends = np.concatenate((mz_gaps, [n]))
    bins = list(zip(mz_starts.tolist(), mz_ends.tolist()))
    nbins = len(bins)

    # The per-bin clustering is independent and order-independent by
    # construction, so bins are fanned out across threads: the dominant cost is
    # the numpy KDE (np.exp over a grid x peaks matrix), which releases the GIL,
    # giving parallelism without multiprocessing's spawn/pickle hazards in the
    # frozen build. Results are collected by bin index and concatenated in the
    # original m/z order, so the feature list is identical to the serial version.
    # `cb` (previously accepted but never called) now reports live per-bin
    # progress + ETA so a long correspondence stage is visibly alive, not hung.
    done = 0
    lock = threading.Lock()
    t0 = time.time()
    last = t0

    def run_bin(bi):
        nonlocal done, last
        a, b = bins[bi]
        gidx = order[a:b]
        clusters = _density_cluster_bin(store, gidx, n_samples, bw_min,
                                        params.min_frac_samples,
                                        capture_mult=params.rt_capture_mult)
        feats = [_feature_from_cluster(store, c) for c in clusters]
        if cb is not None:
            with lock:
                done += 1
                now = time.time()
                if now - last >= 0.5 or done == nbins:
                    frac = done / max(nbins, 1)
                    eta = (now - t0) * (1 - frac) / max(frac, 1e-9)
                    cb(f"  grouping {done:,}/{nbins:,} m/z bins "
                       f"({frac * 100:.0f}%) — ETA {eta:4.0f}s")
                    last = now
        return feats

    results = [None] * nbins
    workers = max(2, (os.cpu_count() or 2))
    with ThreadPoolExecutor(max_workers=workers) as pool:
        for bi, feats in zip(range(nbins), pool.map(run_bin, range(nbins))):
            results[bi] = feats

    features = []
    for feats in results:
        features.extend(feats)
    return features


# ═════════════════════════════════════════════════════════════════════════════
# RT CORRECTION  (anchor features present in most samples get a per-sample
# local-regression correction curve applied to all peaks)
# ═════════════════════════════════════════════════════════════════════════════
def _loess_curve(xs, ys, span=0.2, n_grid=200):
    """
    Local (loess-style) weighted linear regression, matching xcms's own
    PeakGroupsParam(smooth="loess", span=0.2, family="gaussian"): at each of
    `n_grid` points spanning [min(xs), max(xs)], fit a degree-1 regression
    using only the nearest `span`-fraction of anchor points, weighted by a
    tricube kernel on distance (standard loess). Evaluating on a precomputed
    grid (rather than re-fitting per query point) keeps `apply_rt_correction`
    cheap even though it's called once per peak, of which there can be
    hundreds of thousands. Returns (grid_x, grid_y) for np.interp lookup.
    """
    xs = np.asarray(xs); ys = np.asarray(ys)
    order = np.argsort(xs)
    xs, ys = xs[order], ys[order]
    n = len(xs)
    k = min(max(int(np.ceil(span * n)), 3), n)
    lo, hi = float(xs.min()), float(xs.max())
    if hi <= lo:
        return np.array([lo, lo + 1e-6]), np.array([ys.mean(), ys.mean()])

    grid_x = np.linspace(lo, hi, n_grid)
    grid_y = np.empty(n_grid)
    for gi, x0 in enumerate(grid_x):
        d = np.abs(xs - x0)
        idx = np.argpartition(d, k - 1)[:k]
        dk = d[idx]
        maxd = dk.max()
        w = np.ones(k) if maxd <= 0 else (1 - np.clip(dk / maxd, 0, 1) ** 3) ** 3
        X, Y = xs[idx], ys[idx]
        Sw = w.sum(); Sx = (w * X).sum(); Sy = (w * Y).sum()
        Sxx = (w * X * X).sum(); Sxy = (w * X * Y).sum()
        denom = Sw * Sxx - Sx * Sx
        if abs(denom) < 1e-12:
            m, b = 0.0, Sy / max(Sw, 1e-9)
        else:
            m = (Sw * Sxy - Sx * Sy) / denom
            b = (Sy - m * Sx) / Sw
        grid_y[gi] = m * x0 + b
    return grid_x, grid_y


def compute_rt_correction(store, features, n_samples, min_frac_anchor=0.8, span=0.2):
    """
    Pick anchor features present in >= min_frac_anchor of samples, then for
    each sample fit a LOESS-style local regression curve (deviation from the
    cross-sample weighted-mean RT as a function of observed RT) and return a
    per-sample callable correction function.

    Previously this fit ONE global degree-2 polynomial across the entire RT
    range and CLAMPED any query RT to the anchor range before evaluating —
    a single low-order global fit is a poor model for real chromatographic
    drift (which is rarely a clean parabola end-to-end), and clamping instead
    of extrapolating is exactly wrong at the run's edges, where anchor
    density is thinnest and correction is needed most. xcms's own RT
    correction (PeakGroupsParam) uses `smooth="loess", span=0.2` for this
    reason. This now mirrors that: local weighted regression via
    `_loess_curve`, with linear extrapolation (continuing the boundary
    segment's slope) instead of clamping beyond the anchor range.
    """
    anchors = [f for f in features if f["size"] / max(n_samples, 1) >= min_frac_anchor]
    if len(anchors) < 3:
        return {s: (lambda rt: rt) for s in range(n_samples)}

    corrections = {}
    for s in range(n_samples):
        xs, ys = [], []
        for feat in anchors:
            # first member peak of this feature belonging to sample s (in feat
            # order). One-peak-per-sample makes this unique, matching the old
            # [p for p in feat["peaks"] if p["sample"]==s][0].
            found = -1
            for gi in feat["peak_idx"]:
                if int(store.sample[gi]) == s:
                    found = int(gi); break
            if found < 0:
                continue
            observed_rt = float(store.rt[found])
            reference_rt = feat["rt"]
            xs.append(observed_rt)
            ys.append(reference_rt - observed_rt)
        if len(xs) < 3:
            corrections[s] = lambda rt: rt
            continue

        try:
            grid_x, grid_y = _loess_curve(xs, ys, span=span)
        except Exception:
            corrections[s] = lambda rt: rt
            continue

        def make_fn(grid_x, grid_y):
            lo, hi = grid_x[0], grid_x[-1]
            slope_lo = (grid_y[1] - grid_y[0]) / (grid_x[1] - grid_x[0])
            slope_hi = (grid_y[-1] - grid_y[-2]) / (grid_x[-1] - grid_x[-2])

            def fn(rt, grid_x=grid_x, grid_y=grid_y, lo=lo, hi=hi,
                   slope_lo=slope_lo, slope_hi=slope_hi):
                if rt < lo:
                    delta = grid_y[0] + slope_lo * (rt - lo)
                elif rt > hi:
                    delta = grid_y[-1] + slope_hi * (rt - hi)
                else:
                    delta = np.interp(rt, grid_x, grid_y)
                return rt + float(delta)
            return fn

        corrections[s] = make_fn(grid_x, grid_y)
    return corrections


def apply_rt_correction(store, corrections):
    # Mutate rt/rt_min/rt_max columns in place with the SAME per-peak scalar
    # arithmetic as the old dict version: reading each value out as a python float
    # and doing x + (fn(x) - x) reproduces it bit-for-bit (a + (b - a) is NOT
    # generally b in floating point, so this must not be vectorised/simplified).
    ident = lambda rt: rt
    rt = store.rt; rtm = store.rt_min; rtx = store.rt_max; sm = store.sample
    for i in range(store.n):
        fn = corrections.get(int(sm[i]), ident)
        x = float(rt[i])
        delta = fn(x) - x
        rt[i] = x + delta
        rtm[i] = float(rtm[i]) + delta
        rtx[i] = float(rtx[i]) + delta


# ═════════════════════════════════════════════════════════════════════════════
# MS2 ATTACHMENT  — one representative sample's fragmentation per feature
# ═════════════════════════════════════════════════════════════════════════════
def attach_ms2(features, all_ms2_by_sample, grp_params, rt_window_min=0.5):
    """
    For each feature, look up MS2 spectra only in its representative sample
    (the sample with the highest peak height for that feature) whose
    precursor m/z matches within the grouping tolerance and whose RT falls
    within the feature's RT window (+/- rt_window_min). This keeps output
    volume down and avoids redundant identical MS2 across samples.
    """
    # Index each sample's MS2 scans by precursor m/z ONCE, so every feature does
    # a binary search instead of scanning the whole list. The previous linear
    # scan was O(features x MS2 scans) — with ~50k features and ~6k MS2 scans
    # that is ~300M Python-level comparisons, and it scaled directly with the
    # (much larger) feature counts the detection fixes produce.
    index = {}
    for s, lst in all_ms2_by_sample.items():
        if not lst:
            index[s] = ([], [])
            continue
        srt = sorted(lst, key=lambda x: x["prec_mz"])
        index[s] = (srt, [x["prec_mz"] for x in srt])

    for feat in features:
        sample = feat["rep_sample"]
        sorted_ms2, sorted_mz = index.get(sample, ([], []))
        if not sorted_ms2:
            feat["ms2_sample"] = None
            feat["ms2_rt"] = None
            feat["ms2_fragments"] = ""
            continue
        tol = _mz_tol(feat["mz"], grp_params)
        rt_lo = feat["rt_min"] - rt_window_min
        rt_hi = feat["rt_max"] + rt_window_min
        lo = bisect.bisect_left(sorted_mz, feat["mz"] - tol)
        hi = bisect.bisect_right(sorted_mz, feat["mz"] + tol)
        candidates = [
            s for s in sorted_ms2[lo:hi]
            if rt_lo <= s["rt_min"] <= rt_hi
        ]
        if not candidates:
            feat["ms2_sample"] = None
            feat["ms2_rt"] = None
            feat["ms2_fragments"] = ""
            continue
        best = min(candidates, key=lambda s: abs(s["rt_min"] - feat["rt"]))
        feat["ms2_sample"] = sample
        feat["ms2_rt"] = best["rt_min"]
        feat["ms2_fragments"] = summarize_fragments(_get_pairs(best))
    return features


# ═════════════════════════════════════════════════════════════════════════════
# FULL PIPELINE
# ═════════════════════════════════════════════════════════════════════════════
def _detect_worker(args):
    """Detect peaks for one file. MODULE-LEVEL and picklable so it can run in a
    ProcessPoolExecutor worker (real parallelism — detection's mzML parsing is
    GIL-bound, so a thread pool was measured 2.4x SLOWER than serial; processes
    are 2.4x faster than serial with byte-identical peaks). No progress cb (can't
    cross a process boundary); the parent logs per-file on completion. Returns a
    picklable tuple; the caller assigns the sample index."""
    idx, path, cache_dir, peak_params = args
    _noop = lambda *a, **k: None
    try:
        peaks, ms2_scans, kind = extract_peaks_and_ms2(path, cache_dir, peak_params, cb=_noop)
        for p in peaks:
            p["sample"] = idx
        return idx, "OK", peaks, ms2_scans, kind
    except Exception as exc:
        return idx, "ERROR", str(exc), None, None


# ═════════════════════════════════════════════════════════════════════════════
def run_alignment(
    files, out_dir, peak_params, grp_params,
    stop_event, pause_event, log_q,
):
    """
    Background worker. Sends messages to *log_q*:
      ("LOG",  (level, msg))
      ("STAT", (idx, done, features_found, eta_str))
      ("DONE", None)
    """
    log_path = os.path.join(out_dir, "alignment_log.txt")
    # Conversion/extraction cache lives on LOCAL disk, never inside the (often
    # OneDrive-synced) output folder. Derived mzML must not consume cloud quota
    # or sync mid-write — a full or throttled OneDrive during conversion is
    # another route to partial-.mzML corruption. Keyed by the output folder so
    # distinct jobs keep separate caches, exactly as before, just relocated.
    import hashlib as _hashlib
    _job_key = _hashlib.sha1(os.path.abspath(out_dir).encode("utf-8")).hexdigest()[:16]
    _cache_root = os.environ.get("LOCALAPPDATA") or tempfile.gettempdir()
    cache_dir = os.path.join(_cache_root, "VeroMass_Aligner", "raw_convert_cache", _job_key)
    os.makedirs(cache_dir, exist_ok=True)
    start_time = time.time()
    log_lock = threading.Lock()

    def flog(msg, level="INFO"):
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_q.put(("LOG", (level, msg)))
        with log_lock:
            with open(log_path, "a", encoding="utf-8") as fh:
                fh.write(f"[{ts}] [{level}] {msg}\n")

    def cb(msg, level="INFO"):
        flog(msg, level)

    flog("=" * 60)
    flog(f"{TOOL_NAME} v{VERSION} — Started")
    flog(f"Output folder : {out_dir}")
    flog("=" * 60)

    # ── Expand any .zip archives (recursively) into readable MS files ─────────
    n_archives = sum(1 for p in files if p.lower().endswith(ARCHIVE_EXTS))
    if n_archives:
        flog(f"Extracting {n_archives} archive(s)…")
        files = expand_inputs(files, cache_dir, cb)
    else:
        files = [p for p in files if p.lower().endswith(MS_EXTS)]

    total = len(files)
    if total == 0:
        flog("No readable MS files found (after extracting any archives).", "WARN")
        log_q.put(("DONE", None))
        return
    flog(f"Total MS files: {total}")

    sample_names = [os.path.basename(p) for p in files]

    # ── Pre-convert RAW files CONCURRENTLY ────────────────────────────────────
    # RAW->mzML conversion is an external ThermoRawFileParser subprocess (~30s
    # each) that does NOT hold the Python GIL, so running conversions in
    # parallel threads turns N sequential conversions into ~one conversion's
    # wall-time. (Peak detection is left serial below — it's CPU-bound Python
    # and the GIL makes threads there slower than serial, so parallelizing it
    # would need processes; the vectorized detector is already ~1s/file.)
    raw_files = [p for p in files if p.lower().endswith(".raw")]
    if raw_files and not stop_event.is_set():
        n_conv = min(len(raw_files), max(2, (os.cpu_count() or 2)), 6)
        flog(f"Pre-converting {len(raw_files)} RAW file(s), {n_conv} at a time…")

        def _convert_one(path):
            if stop_event.is_set():
                return
            try:
                convert_raw(path, cache_dir, cb)
            except Exception as exc:
                flog(f"  [{os.path.basename(path)}] conversion ERROR: {exc}", "ERROR")

        with ThreadPoolExecutor(max_workers=n_conv) as pool:
            list(pool.map(_convert_one, raw_files))

    # ── Peak detection: PROCESS POOL across files, with a serial fallback ─────
    # Each file is independent. Detection is dominated by mzML PARSING, which is
    # GIL-bound Python — so a thread pool is measured 2.4x SLOWER than serial
    # (threads just time-slice the GIL and add overhead). A PROCESS pool gives
    # true parallelism: ~2.4x faster than serial, ~5.7x faster than threads, with
    # byte-identical peaks. If a process pool can't spawn (e.g. a onefile frozen
    # build that mis-bootstraps its children), we fall back to SERIAL — never to
    # threads — so the result is always at least as fast as sequential detection.
    # Peaks are converted to columnar PeakStore CHUNKS as each file completes and
    # the per-file dicts are dropped immediately, so only one file's dicts are ever
    # alive while the columnar store (70 B/peak) grows — this is what keeps a
    # 300-file run inside ~2.4 GB instead of ~17 GB of dicts. Chunks are
    # concatenated once, after detection, preserving order for the stable m/z sort.
    peak_chunks = []
    all_ms2_by_sample = {}
    n_ok = 0
    n_done = 0
    n_peaks_total = 0
    n_workers = max(2, min((os.cpu_count() or 2), 8))
    tasks = [(idx, path, cache_dir, peak_params) for idx, path in enumerate(files)]

    def _collect(rec):
        nonlocal n_ok, n_done, n_peaks_total
        idx, status, a, b, c = rec
        n_done += 1
        name = os.path.basename(files[idx])
        if status == "ERROR":
            flog(f"[{idx + 1}/{total}] {name}  ERROR: {a}", "ERROR")
        else:
            peak_chunks.append(PeakStore.from_dicts(a))   # a=peaks (dicts freed after)
            n_peaks_total += len(a)
            all_ms2_by_sample[idx] = b                    # b=ms2_scans, c=kind
            n_ok += 1
            flog(f"[{idx + 1}/{total}] {name}: {c}; {len(a)} peaks.", "OK")
        log_q.put(("STAT", (n_done, n_ok, n_peaks_total,
                            _calc_eta(start_time, n_done, total - n_done))))

    detected = False
    if not stop_event.is_set():
        flog(f"Detecting peaks in {total} file(s), {n_workers}-way parallel…")
        try:
            from concurrent.futures import ProcessPoolExecutor
            with ProcessPoolExecutor(max_workers=n_workers) as pool:
                for rec in pool.map(_detect_worker, tasks):
                    if stop_event.is_set():
                        break
                    _collect(rec)
            detected = True
        except Exception as exc:
            flog(f"  Parallel detection unavailable ({exc}); using serial.", "WARN")
            peak_chunks.clear(); all_ms2_by_sample.clear()
            n_ok = 0; n_done = 0; n_peaks_total = 0

    if not detected and not stop_event.is_set():
        flog(f"Detecting peaks in {total} file(s), serial…")
        for task in tasks:
            if stop_event.is_set():
                break
            while not pause_event.is_set():
                if stop_event.is_set():
                    break
                time.sleep(0.2)
            if stop_event.is_set():
                break
            _collect(_detect_worker(task))

    if stop_event.is_set():
        flog("Stopped by user.", "WARN")

    if stop_event.is_set() or n_ok == 0:
        flog("No samples processed — aborting before grouping.", "WARN")
        log_q.put(("DONE", None))
        return

    # Denominator for the min-fraction filter must be the number of samples that
    # ACTUALLY produced peaks, not the total queued. Using `total` meant a partial
    # run (e.g. 3 of 14 converted) demanded features appear in a fraction of 14,
    # which silently guaranteed an empty table. Fail loud when samples were lost
    # so a degraded run is obvious rather than quietly wrong.
    n_samples = n_ok
    if n_ok < total:
        flog(f"WARNING: only {n_ok} of {total} sample(s) produced peaks — "
             f"{total - n_ok} failed (see ERROR lines above). Grouping now uses "
             f"a denominator of {n_ok}; review the failures before trusting "
             f"these results.", "WARN")
    # Fold the per-file chunks into one columnar store; drop the chunk list so the
    # only large object from here on is the store itself.
    store = PeakStore.concat(peak_chunks)
    peak_chunks.clear()
    flog(f"Peak picking complete: {store.n} peaks across {n_samples} sample(s) "
         f"(columnar store {store.nbytes() / 1e6:.0f} MB).")

    flog("Correspondence — grouping peaks across samples…")
    features = group_peaks_density(store, n_samples, grp_params, cb=cb)
    flog(f"  {len(features)} feature(s) after initial grouping.")

    flog("Computing per-sample RT correction from anchor features…")
    corrections = compute_rt_correction(store, features, n_samples)
    apply_rt_correction(store, corrections)

    flog("Re-grouping on RT-corrected peaks…")
    features = group_peaks_density(store, n_samples, grp_params, cb=cb)
    flog(f"  {len(features)} final feature(s) after RT-corrected re-grouping.")

    flog("Attaching representative MS2 fragmentation per feature…")
    features = attach_ms2(features, all_ms2_by_sample, grp_params)
    n_with_ms2 = sum(1 for f in features if f["ms2_fragments"])
    flog(f"  {n_with_ms2}/{len(features)} feature(s) matched to an MS2 spectrum.")

    flog("Writing Excel output…")
    out_path, peaks_csv = write_feature_table(store, features, sample_names, out_dir)
    flog(f"Output written: {out_path}", "OK")
    if peaks_csv:
        flog(f"Per-peak table written separately (too large for a worksheet): {peaks_csv}", "OK")

    # Atomic completion marker for the desktop Bridge's watched-folder mode
    # (see veromass-bridge/watch.py): write to a temp name in the same
    # directory, then os.replace — a reader can never observe a half-written
    # marker, only "absent" or "fully present".
    ready_tmp = os.path.join(out_dir, ".ready.tmp")
    with open(ready_tmp, "w", encoding="utf-8") as fh:
        fh.write(f"{TOOL_NAME} v{VERSION}\n")
    os.replace(ready_tmp, os.path.join(out_dir, ".ready"))

    flog("=" * 60 + f"\nFinished  |  samples={n_samples}  features={len(features)}", "OK")
    log_q.put(("STAT", (total, n_ok, len(features), "—")))
    log_q.put(("DONE", None))


def _calc_eta(start_time, done, remaining):
    if done <= 0:
        return "—"
    avg = (time.time() - start_time) / done
    secs = int(avg * remaining)
    return str(datetime.timedelta(seconds=secs))


def scan_input_files(folder, recurse):
    result = []
    if recurse:
        for root, _dirs, fs in os.walk(folder):
            for f in fs:
                if f.lower().endswith(SUPPORTED_EXTS):
                    result.append(os.path.join(root, f))
    else:
        for f in sorted(os.listdir(folder)):
            if f.lower().endswith(SUPPORTED_EXTS):
                result.append(os.path.join(folder, f))
    return sorted(result)


def expand_inputs(paths, work_dir, cb=None, _depth=0):
    """Return a flat list of directly-readable MS files, extracting any .zip
    archives (recursively — supports nested zips like a dataset zip whose
    entries are per-sample zips each containing one mzML). Extracted files land
    under work_dir/unzipped/ and are re-scanned for MS files and further zips."""
    out = []
    for p in paths:
        ext = Path(p).suffix.lower()
        if ext in MS_EXTS:
            out.append(p)
        elif ext in ARCHIVE_EXTS and _depth < 6:
            dest = os.path.join(work_dir, "unzipped", f"{Path(p).stem}_{_depth}")
            os.makedirs(dest, exist_ok=True)
            try:
                with zipfile.ZipFile(p) as z:
                    z.extractall(dest)
            except Exception as exc:
                if cb:
                    cb(f"  Could not open archive {Path(p).name}: {exc}", "ERROR")
                continue
            inner = []
            for root, _dirs, fs in os.walk(dest):
                for f in fs:
                    if f.lower().endswith(SUPPORTED_EXTS):
                        inner.append(os.path.join(root, f))
            out.extend(expand_inputs(sorted(inner), work_dir, cb, _depth + 1))
    # de-dup while preserving order
    seen = set()
    uniq = []
    for p in out:
        if p not in seen:
            seen.add(p)
            uniq.append(p)
    return uniq


def extract_peaks_and_ms2(path, cache_dir, peak_params, cb=None):
    """Produce (peaks, ms2_scans, kind_label) for one file, dispatching on type.
    MGF spectra are used directly as peaks (no chromatographic detection);
    mzML/mzXML/RAW go through full peak detection."""
    ext = Path(path).suffix.lower()
    if ext == ".mgf":
        spectra = read_mgf_spectra(path, cb)
        return peaks_from_mgf(spectra), spectra, f"{len(spectra):,} MGF spectra"
    if ext == ".mzml":
        ms1, ms2 = read_ms_scans_mzml(path, cb)
    elif ext == ".mzxml":
        ms1, ms2 = read_ms_scans_mzxml(path, cb)
    elif ext == ".raw":
        mzml_path = convert_raw(path, cache_dir, cb or (lambda *a: None))
        ms1, ms2 = read_ms_scans_mzml(mzml_path, cb)
    else:
        raise ValueError(f"Unsupported file type: {ext}")
    peaks = detect_chrom_peaks(ms1, peak_params, cb=cb)
    return peaks, ms2, f"{len(ms1)} MS1 / {len(ms2)} MS2 scans"


# ═════════════════════════════════════════════════════════════════════════════
# EXPORT
# ═════════════════════════════════════════════════════════════════════════════
FEATURE_COLUMNS = [
    "feature", "Size", "Charge", "Mass", "m.z", "RT", "Base.Peak",
    "m.z.Width", "RT.Height", "m.z.Min", "m.z.Max", "RT.Min", "RT.Max",
    "Shape.Distance", "MS2.Sample", "MS2.RT", "MS2.Fragments",
]

PEAK_COLUMNS = ["feature", "sample", "m.z", "RT", "Height", "Area", "S.N"]

# Above this many per-peak rows the Peaks table is written as a companion CSV
# instead of a worksheet — writing 231k rows via openpyxl costs ~73 s vs ~2 s
# to CSV, and that write was a third of total runtime on real 2-sample studies.
PEAKS_XLSX_MAX_ROWS = 100_000


def write_feature_table(store, features, sample_names, out_dir):
    """
    Writes three sheets:
      Features    — one row per aligned feature (cross-sample summary + MS2)
      Peaks       — one row per detected peak per sample per feature: the
                    full list of detected m/z, RT, and intensity (height/area)
                    across every chromatogram used in the alignment
      Intensities — wide per-sample intensity matrix (one column per sample),
                    convenient for stats tools that expect a matrix layout
    """
    out_path = os.path.join(out_dir, "aligned_features.xlsx")
    feature_rows = []
    peak_rows = []
    intensity_rows = []

    for i, feat in enumerate(sorted(features, key=lambda f: f["rt"]), 1):
        feature_id = f"Feature_{i:06d}"
        mass = feat["mz"] - 1.007276470  # assume singly-charged [M+H]+ by default
        feature_rows.append({
            "feature": feature_id,
            "Size": feat["size"],
            "Charge": 1,
            "Mass": round(mass, 9),
            "m.z": round(feat["mz"], 9),
            "RT": round(feat["rt"], 10),
            "Base.Peak": int(round(feat["base_peak"])),
            "m.z.Width": round(feat["mz_max"] - feat["mz_min"], 11),
            "RT.Height": round(feat["rt_max"] - feat["rt_min"], 12),
            "m.z.Min": round(feat["mz_min"], 9),
            "m.z.Max": round(feat["mz_max"], 9),
            "RT.Min": round(feat["rt_min"], 10),
            "RT.Max": round(feat["rt_max"], 10),
            "Shape.Distance": feat["shape_distance"],
            "MS2.Sample": sample_names[feat["ms2_sample"]] if feat.get("ms2_sample") is not None else "",
            "MS2.RT": round(feat["ms2_rt"], 4) if feat.get("ms2_rt") is not None else "",
            "MS2.Fragments": feat.get("ms2_fragments", ""),
        })

        by_sample = {}
        for gi in feat["peak_idx"]:
            gi = int(gi)
            s = int(store.sample[gi])
            h = float(store.height[gi])          # float32 store -> exact float64
            by_sample.setdefault(s, []).append(h)
            peak_rows.append({
                "feature": feature_id,
                "sample": sample_names[s],
                "m.z": round(float(store.mz[gi]), 9),
                "RT": round(float(store.rt[gi]), 6),
                "Height": round(h, 2),
                "Area": round(float(store.area[gi]), 2),
                "S.N": round(float(store.snr[gi]), 2),
            })

        intens_row = {"feature": feature_id}
        for s, name in enumerate(sample_names):
            vals = by_sample.get(s)
            intens_row[name] = max(vals) if vals else None
        intensity_rows.append(intens_row)

    df_feat = pd.DataFrame(feature_rows, columns=FEATURE_COLUMNS)
    df_peaks = pd.DataFrame(peak_rows, columns=PEAK_COLUMNS)
    df_int = pd.DataFrame(intensity_rows, columns=["feature"] + sample_names)

    def _col_widths(df):
        # Compute widths from the header + a sample of rows (not every cell —
        # scanning all cells via ws.columns is O(rows×cols) and dominates the
        # write for large Peaks sheets). A 500-row sample is plenty for sizing.
        widths = []
        sample = df.head(500)
        for col in df.columns:
            header_len = len(str(col))
            if len(sample):
                body_len = sample[col].astype(str).str.len().max()
                body_len = 0 if pd.isna(body_len) else int(body_len)
            else:
                body_len = 0
            widths.append(min(max(header_len, body_len) + 2, 40))
        return widths

    # The per-peak table is the only one that gets genuinely large (one row per
    # detected peak per sample). openpyxl costs ~73 s for 231k rows, xlsxwriter
    # ~50 s, CSV ~2 s — so past a threshold it goes to a companion CSV instead
    # of a worksheet. Excel opens the CSV directly and stats tools prefer it.
    sheets = [("Features", df_feat), ("Peaks", df_peaks), ("Intensities", df_int)]
    peaks_csv_path = None
    if len(df_peaks) > PEAKS_XLSX_MAX_ROWS:
        peaks_csv_path = os.path.join(out_dir, "aligned_peaks.csv")
        df_peaks.to_csv(peaks_csv_path, index=False)
        sheets = [("Features", df_feat), ("Intensities", df_int)]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            hdr_fill = PatternFill("solid", fgColor="1C2640")
            hdr_font = Font(bold=True, color="FFFFFF")
            border = Border(bottom=Side(style="medium", color="00E5C0"))
            for cell in ws[1]:
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 20
            for ci, width in enumerate(_col_widths(df), 1):
                ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = width

    return out_path, peaks_csv_path


# ═════════════════════════════════════════════════════════════════════════════
# GUI
# ═════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        # When launched via "Process locally" (veromass-bridge/launcher.py),
        # these env vars carry the Workbench/Job name AND id the scientist
        # already sees in app.veromass.com — shown here so the desktop run
        # is visibly the same job, not just silently linked by a UUID the
        # commit path already guarantees underneath. Unset on a normal
        # manual launch (`python VeroMass_Aligner.py`) — title/UI are then
        # unchanged from before.
        self._linked_workbench_name = os.environ.get("VEROMASS_WORKBENCH_NAME") or None
        self._linked_job_name = os.environ.get("VEROMASS_JOB_NAME") or None
        self._linked_job_id = os.environ.get("VEROMASS_JOB_ID") or None

        title = f"{TOOL_NAME}  v{VERSION}"
        if self._linked_job_name or self._linked_workbench_name:
            title += f"  —  {self._linked_workbench_name or '?'} / {self._linked_job_name or '(untitled job)'}"
        self.title(title)
        # A fixed "1000x780" put the bottom button bar (Start/Pause/
        # Reset/Open Output Folder) off-screen on smaller/lower-resolution
        # displays — the window simply extended past the visible desktop
        # area (behind the taskbar or off the bottom edge entirely), and a
        # user had to manually drag-resize or maximize to ever see those
        # buttons. Size relative to the REAL screen instead, capped at the
        # old 1000x780 on large displays, and center it — the whole window
        # (buttons included) is now always inside the visible screen at
        # launch, on any display, with zero manual resizing needed.
        screen_w, screen_h = self.winfo_screenwidth(), self.winfo_screenheight()
        win_w = min(1000, int(screen_w * 0.92))
        win_h = min(780, int(screen_h * 0.88))
        x, y = (screen_w - win_w) // 2, (screen_h - win_h) // 2
        self.geometry(f"{win_w}x{win_h}+{x}+{y}")
        self.minsize(min(860, win_w), min(640, win_h))
        self.configure(bg=C_BG)

        self._stop_ev = threading.Event()
        self._pause_ev = threading.Event()
        self._pause_ev.set()
        self._log_q = queue.Queue()
        self._thread = None
        self._total = 0

        self._build_ui()
        if os.environ.get("VEROMASS_OUTPUT_DIR"):
            self._log("INFO", f"Output folder auto-set for Bridge pickup: {self._outfolder_var.get()}")
        self.after(150, self._poll)

    # ── UI construction ───────────────────────────────────────────────────────
    def _build_ui(self):
        tf = tk.Frame(self, bg=C_BG)
        tf.pack(fill="x", padx=24, pady=(18, 2))
        tk.Label(tf, text=TOOL_NAME, bg=C_BG, fg=C_TEAL,
                 font=("Segoe UI", 17, "bold")).pack(side="left")
        tk.Label(tf, text=f" v{VERSION}  ·  VeroMass / MoleculeID Platform  ·  Standalone Utility",
                 bg=C_BG, fg=C_DIM, font=("Segoe UI", 9)).pack(side="left", padx=8)

        if self._linked_job_name or self._linked_workbench_name:
            short_id = f"…{self._linked_job_id[-8:]}" if self._linked_job_id else ""
            tk.Label(
                tf,
                text=f"  Linked to: {self._linked_workbench_name or '?'} → "
                     f"{self._linked_job_name or '(untitled job)'} ({short_id})",
                bg=C_BG, fg=C_TEAL, font=("Segoe UI", 9, "bold"),
            ).pack(side="left", padx=(12, 0))

        # ── Input ──
        inp = self._section("  Input  ")
        inp.pack(fill="x", padx=20, pady=(8, 0))

        self._folder_var = tk.StringVar()
        # When launched via "Process locally", veromass-bridge sets
        # VEROMASS_OUTPUT_DIR to a per-job subfolder of its own watched
        # folder (watch.py's DEFAULT_DIR) — pre-filling it here means the
        # scientist never has to know/type that path for the Bridge to
        # auto-pick-up the finished run. A manual launch (env var unset)
        # is unchanged — empty, same as before.
        self._outfolder_var = tk.StringVar(value=os.environ.get("VEROMASS_OUTPUT_DIR", ""))
        self._recurse_var = tk.BooleanVar(value=True)

        self._path_row(inp, "Runs folder:", self._folder_var, self._browse_in)
        self._path_row(inp, "Output folder:", self._outfolder_var, self._browse_out)

        opt = tk.Frame(inp, bg=C_BG)
        opt.pack(fill="x", padx=12, pady=(4, 10))
        self._chk(opt, "Scan subfolders recursively", self._recurse_var)
        tk.Label(opt, text="  Accepts: .mzML  .mzXML  .raw (Thermo)  .mgf  .zip",
                 bg=C_BG, fg=C_DIM, font=("Segoe UI", 8)).pack(side="left", padx=(20, 0))

        # ── Parameters ──
        par = self._section("  Parameters  ")
        par.pack(fill="x", padx=20, pady=(8, 0))
        prow = tk.Frame(par, bg=C_BG)
        prow.pack(fill="x", padx=12, pady=(8, 10))

        # Seed every field FROM the parameter defaults rather than repeating
        # literals here. These fields override the dataclass on every run, so a
        # stale literal silently reverts a tuning fix — a leftover "5" for the
        # minimum peak width cost ~20x the detected peak count (5,636 vs
        # 111,889 on the same file) while the code default already said 1.
        _pd = PeakDetectionParams()
        _gd = GroupingParams()

        def _fmt(v):
            return str(int(v)) if float(v) == int(v) else str(v)

        self._ppm_var = tk.StringVar(value=_fmt(_pd.ppm))
        self._minw_var = tk.StringVar(value=_fmt(_pd.peak_min_sec))
        self._maxw_var = tk.StringVar(value=_fmt(_pd.peak_max_sec))
        self._snr_var = tk.StringVar(value=_fmt(_pd.snr_thresh))
        self._noise_var = tk.StringVar(value=_fmt(_pd.noise))
        self._grp_ppm_var = tk.StringVar(value=_fmt(_gd.mz_ppm))
        self._rtbw_var = tk.StringVar(value=_fmt(_gd.rt_bw_sec))
        self._minfrac_var = tk.StringVar(value=_fmt(_gd.min_frac_samples))

        specs = [
            ("ppm", self._ppm_var), ("peak min (s)", self._minw_var),
            ("peak max (s)", self._maxw_var), ("S/N", self._snr_var),
            ("noise", self._noise_var), ("group ppm", self._grp_ppm_var),
            ("RT bandwidth (s)", self._rtbw_var), ("min sample frac", self._minfrac_var),
        ]
        for label, var in specs:
            cell = tk.Frame(prow, bg=C_BG)
            cell.pack(side="left", padx=8)
            tk.Label(cell, text=label, bg=C_BG, fg=C_SUB, font=("Segoe UI", 8)).pack()
            tk.Entry(cell, textvariable=var, width=8, bg=C_PANEL, fg=C_FG,
                     insertbackground=C_FG, relief="flat", justify="center",
                     font=("Consolas", 9)).pack()

        # ── Progress ──
        prg = self._section("  Progress  ")
        prg.pack(fill="x", padx=20, pady=(8, 0))

        cards = tk.Frame(prg, bg=C_BG)
        cards.pack(fill="x", padx=12, pady=(10, 4))
        self._sv_processed = self._stat_card(cards, "Files done", "0", C_TEAL)
        self._sv_total = self._stat_card(cards, "Total files", "0", C_FG)
        self._sv_peaks = self._stat_card(cards, "Peaks/Features", "0", C_PURP)
        self._sv_eta = self._stat_card(cards, "ETA", "—", C_GREEN)

        pb_f = tk.Frame(prg, bg=C_BG)
        pb_f.pack(fill="x", padx=12, pady=(4, 2))
        sty = ttk.Style()
        sty.theme_use("clam")
        sty.configure("T.Horizontal.TProgressbar", troughcolor=C_PANEL, background=C_TEAL,
                       darkcolor=C_TEAL, lightcolor=C_TEAL, bordercolor=C_PANEL)
        self._pb = ttk.Progressbar(pb_f, orient="horizontal", length=100,
                                    mode="determinate", style="T.Horizontal.TProgressbar")
        self._pb.pack(fill="x", expand=True)
        self._pb_lbl = tk.Label(pb_f, text="0 / 0  (0%)", bg=C_BG, fg=C_SUB, font=("Segoe UI", 8))
        self._pb_lbl.pack(pady=(2, 6))

        # ── Buttons ── packed BEFORE the Log section, anchored to the
        # window's bottom edge (side="bottom"). Tkinter's pack manager
        # allocates space to already-packed widgets first; the Log section
        # below has fill="both"/expand=True, which — when packed FIRST, as
        # it used to be — claims all available space at layout time and
        # pushes whatever is packed after it (this button row) out of the
        # visible window entirely on any display too short to fit
        # everything at full size. Packing the buttons first (and anchoring
        # them to the bottom rather than relying on top-down stacking order
        # alone) guarantees they always get their own space; Log then only
        # ever fills whatever room is actually left over, shrinking or
        # scrolling instead of hiding the controls.
        btn_f = tk.Frame(self, bg=C_BG)
        btn_f.pack(side="bottom", fill="x", padx=20, pady=12)
        self._btn_start = self._btn(btn_f, "▶  Start", bg=C_TEAL, fg="#000000", bold=True, cmd=self._start)
        self._btn_start.pack(side="left", padx=(0, 8))
        self._btn_pause = self._btn(btn_f, "⏸  Pause", bg=C_BORDER, fg=C_FG, cmd=self._toggle_pause, state="disabled")
        self._btn_pause.pack(side="left", padx=(0, 8))
        self._btn_stop = self._btn(btn_f, "⏹  Stop", bg="#2A1520", fg=C_RED, cmd=self._stop, state="disabled")
        self._btn_stop.pack(side="left")
        self._btn_reset = self._btn(btn_f, "↺  Reset", bg=C_BORDER, fg=C_AMB, cmd=self._reset)
        self._btn_reset.pack(side="left", padx=(8, 0))
        self._btn_open = self._btn(btn_f, "\U0001F4C2  Open Output Folder", bg=C_BORDER, fg=C_FG, cmd=self._open_out)
        self._btn_open.pack(side="right")

        # ── Log ──
        log_sec = self._section("  Log  ")
        log_sec.pack(fill="both", expand=True, padx=20, pady=(8, 0))
        self._log_w = scrolledtext.ScrolledText(
            log_sec, bg="#080E1A", fg=C_SUB, font=("Consolas", 8),
            relief="flat", state="disabled", wrap="word",
        )
        self._log_w.pack(fill="both", expand=True, padx=6, pady=6)
        for tag, col in (("INFO", C_SUB), ("OK", C_TEAL), ("WARN", C_AMB), ("ERROR", C_RED)):
            self._log_w.tag_config(tag, foreground=col)

    # ── Widget helpers ────────────────────────────────────────────────────────
    def _section(self, title):
        return tk.LabelFrame(self, text=title, bg=C_BG, fg=C_TEAL, font=("Segoe UI", 9, "bold"),
                              bd=1, relief="groove", highlightbackground=C_BORDER)

    def _path_row(self, parent, label, var, cmd):
        row = tk.Frame(parent, bg=C_BG)
        row.pack(fill="x", padx=12, pady=(6, 2))
        tk.Label(row, text=label, bg=C_BG, fg=C_FG, width=15, font=("Segoe UI", 9), anchor="w").pack(side="left")
        tk.Entry(row, textvariable=var, bg=C_PANEL, fg=C_FG, insertbackground=C_FG,
                  relief="flat", bd=4, font=("Consolas", 9)).pack(side="left", fill="x", expand=True, padx=(0, 8))
        tk.Button(row, text="Browse…", bg=C_BORDER, fg=C_FG, relief="flat", padx=10,
                  cursor="hand2", command=cmd).pack(side="left")

    def _chk(self, parent, text, var, padx=(0, 0)):
        tk.Checkbutton(parent, text=text, variable=var, bg=C_BG, fg=C_FG, selectcolor=C_PANEL,
                        activebackground=C_BG, activeforeground=C_TEAL, font=("Segoe UI", 9)).pack(side="left", padx=padx)

    def _stat_card(self, parent, label, value, color):
        frame = tk.Frame(parent, bg=C_PANEL, bd=0)
        frame.pack(side="left", padx=6, ipadx=14, ipady=8)
        lbl = tk.Label(frame, text=value, bg=C_PANEL, fg=color, font=("Segoe UI", 20, "bold"))
        lbl.pack()
        tk.Label(frame, text=label, bg=C_PANEL, fg=C_DIM, font=("Segoe UI", 8)).pack()
        return lbl

    def _btn(self, parent, text, bg, fg, cmd, state="normal", bold=False):
        return tk.Button(parent, text=text, bg=bg, fg=fg, font=("Segoe UI", 9, "bold" if bold else "normal"),
                          relief="flat", padx=16, pady=6, cursor="hand2", command=cmd, state=state)

    # ── Browse callbacks ──────────────────────────────────────────────────────
    def _browse_in(self):
        d = filedialog.askdirectory(title="Select folder containing mzML / mzXML / RAW / MGF / ZIP files")
        if d:
            self._folder_var.set(d)
            if not self._outfolder_var.get():
                self._outfolder_var.set(os.path.join(d, OUTPUT_SUBDIR))

    def _browse_out(self):
        d = filedialog.askdirectory(title="Select output folder")
        if d:
            self._outfolder_var.set(d)

    def _open_out(self):
        p = self._outfolder_var.get()
        if p and os.path.isdir(p):
            os.startfile(p)
        else:
            self._log("WARN", "Output folder does not exist yet.")

    # ── Start / Pause / Stop ──────────────────────────────────────────────────
    def _start(self):
        folder = self._folder_var.get().strip()
        out_dir = self._outfolder_var.get().strip()

        if not folder or not os.path.isdir(folder):
            self._log("ERROR", "Please select a valid runs folder.")
            return
        if not out_dir:
            out_dir = os.path.join(folder, OUTPUT_SUBDIR)
            self._outfolder_var.set(out_dir)
        os.makedirs(out_dir, exist_ok=True)

        files = scan_input_files(folder, self._recurse_var.get())
        if not files:
            self._log("WARN", "No mzML/mzXML/RAW/MGF/ZIP files found.")
            return

        try:
            peak_params = PeakDetectionParams(
                ppm=float(self._ppm_var.get()),
                peak_min_sec=float(self._minw_var.get()),
                peak_max_sec=float(self._maxw_var.get()),
                snr_thresh=float(self._snr_var.get()),
                noise=float(self._noise_var.get()),
            )
            grp_params = GroupingParams(
                mz_ppm=float(self._grp_ppm_var.get()),
                rt_bw_sec=float(self._rtbw_var.get()),
                min_frac_samples=float(self._minfrac_var.get()),
            )
        except ValueError:
            self._log("ERROR", "One or more parameters is not a valid number.")
            return

        self._total = len(files)
        for sv, val in ((self._sv_processed, "0"), (self._sv_total, str(self._total)),
                         (self._sv_peaks, "0"), (self._sv_eta, "—")):
            sv["text"] = val
        self._pb["value"] = 0
        self._pb_lbl["text"] = f"0 / {self._total}  (0%)"

        self._stop_ev.clear()
        self._pause_ev.set()
        self._btn_start["state"] = "disabled"
        self._btn_pause["state"] = "normal"
        self._btn_stop["state"] = "normal"

        self._log("INFO", f"Queued {self._total} file(s).  Output -> {out_dir}")

        self._thread = threading.Thread(
            target=run_alignment,
            args=(files, out_dir, peak_params, grp_params,
                  self._stop_ev, self._pause_ev, self._log_q),
            daemon=True,
        )
        self._thread.start()

    def _toggle_pause(self):
        if self._pause_ev.is_set():
            self._pause_ev.clear()
            self._btn_pause["text"] = "▶  Resume"
            self._log("WARN", "Paused.  Click Resume to continue.")
        else:
            self._pause_ev.set()
            self._btn_pause["text"] = "⏸  Pause"
            self._log("INFO", "Resumed.")

    def _stop(self):
        self._stop_ev.set()
        self._pause_ev.set()
        self._log("WARN", "Stop requested — finishing current file…")

    def _reset(self):
        """Wipe out the CURRENT job's in-progress/finished state — log,
        progress bars, stat cards, and the run thread — so a fresh run can
        start clean. Does NOT touch the chosen runs/output folder paths,
        parameter fields, any file already written to disk, or anything
        outside this one GUI's own in-memory state (no other job, workbench,
        or committed cloud result is reachable from here)."""
        if self._thread is not None and self._thread.is_alive():
            if not messagebox.askyesno(
                "Reset current job",
                "A run is currently in progress. Stop it and reset?",
            ):
                return
            self._log("WARN", "Reset requested — stopping current run first…")
            self._stop_ev.set()
            self._pause_ev.set()
            self._btn_reset["state"] = "disabled"
            self.after(200, self._wait_for_stop_then_reset)
            return
        self._do_reset()

    def _wait_for_stop_then_reset(self):
        if self._thread is not None and self._thread.is_alive():
            self.after(200, self._wait_for_stop_then_reset)
            return
        self._btn_reset["state"] = "normal"
        self._do_reset()

    def _do_reset(self):
        self._thread = None
        self._total = 0
        self._stop_ev.clear()
        self._pause_ev.set()

        for sv in (self._sv_processed, self._sv_total, self._sv_peaks):
            sv["text"] = "0"
        self._sv_eta["text"] = "—"
        self._pb["value"] = 0
        self._pb_lbl["text"] = "0 / 0  (0%)"

        self._log_w.configure(state="normal")
        self._log_w.delete("1.0", "end")
        self._log_w.configure(state="disabled")

        self._btn_start["state"] = "normal"
        self._btn_pause["state"] = "disabled"
        self._btn_pause["text"] = "⏸  Pause"
        self._btn_stop["state"] = "disabled"

        self._log("INFO", "Job reset. Ready for a new run.")

    # ── Queue polling ─────────────────────────────────────────────────────────
    def _poll(self):
        try:
            while True:
                msg_type, payload = self._log_q.get_nowait()
                if msg_type == "LOG":
                    self._log(*payload)
                elif msg_type == "STAT":
                    idx, done, peaks, eta = payload
                    pct = int(idx / max(self._total, 1) * 100)
                    self._pb["value"] = pct
                    self._pb_lbl["text"] = f"{idx} / {self._total}  ({pct}%)"
                    self._sv_processed["text"] = str(done)
                    self._sv_peaks["text"] = f"{peaks:,}"
                    self._sv_eta["text"] = eta
                elif msg_type == "DONE":
                    self._pb["value"] = 100
                    self._btn_start["state"] = "normal"
                    self._btn_pause["state"] = "disabled"
                    self._btn_stop["state"] = "disabled"
                    self._btn_pause["text"] = "⏸  Pause"
        except queue.Empty:
            pass
        self.after(150, self._poll)

    def _log(self, level, msg):
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        tag = level if level in ("INFO", "OK", "WARN", "ERROR") else "INFO"
        self._log_w.configure(state="normal")
        self._log_w.insert("end", f"[{ts}] {msg}\n", tag)
        self._log_w.configure(state="disabled")
        self._log_w.see("end")


# ═════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ═════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    # Required so ProcessPoolExecutor workers (used for peak detection) spawn
    # correctly under a frozen PyInstaller build instead of re-launching the GUI.
    # A no-op when running from source; must be the first thing in __main__.
    import multiprocessing
    multiprocessing.freeze_support()
    app = App()
    app.mainloop()
